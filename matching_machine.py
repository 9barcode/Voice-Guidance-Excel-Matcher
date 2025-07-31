import pandas as pd
import openpyxl


csv_file_path = input("result.csv 파일 경로 입력 : ")
excel_file_path = input("excel 결과 문서 경로 입력 : ")

# 검색할 값이 있는 CSV와 비교할 엑셀 파일 경로
csv_file = f"{csv_file_path}\\result.csv"
excel_file = f"{excel_file_path}\\Polish_hjkwon.xlsx"

# CSV에서 검색할 값 리스트 불러오기
search_terms = pd.read_csv(csv_file, header=None)[0].tolist()

# B.xlsx 불러오기
wb = openpyxl.load_workbook(excel_file, data_only=True)



# 검색 수행
for term in search_terms:
    found = False
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value) == str(term):  # 정확히 일치하는 값만
                    print(f"[{term}] => {sheet_name}!{cell.coordinate}")
                    found = True
    if not found:
        print(f"[{term}] => 찾지 못함")
